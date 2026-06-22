"""
Management command: import_data
Importa projectos e programações a partir de ficheiros CSV, Excel (.xlsx) ou JSON.

Format attendu (colonnes / clés) :
  code, title, ministry, sector, pillar, financier, status,
  don_2026, emp_2026, state_2026,
  don_2027, emp_2027, state_2027,
  don_2028, emp_2028, state_2028,
  don_2029, emp_2029, state_2029,
  don_2030, emp_2030, state_2030

Colonnes optionnelles :
  description, start_date, end_date, region, is_national

Résolution des FK par nom/sigle/code :
  - ministry  → Ministry.short_name ou Ministry.name
  - sector    → Sector.code ou Sector.label
  - pillar    → Pillar.code ou Pillar.label
  - financier → Financier.short_name ou Financier.name

Uso:
  python manage.py import_data --file dados.csv
  python manage.py import_data --file dados.xlsx
  python manage.py import_data --file dados.json
  python manage.py import_data --file dados.csv --dry-run
  python manage.py import_data --file dados.csv --clear
"""

import csv
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from sigip.models import (
    Pillar, Sector, Ministry, Financier, FinancierType,
    Project, ProjectStatus, WorkflowStatus, AnnualProgramming,
)
from core.models import Tenant, Region

ANOS = [2026, 2027, 2028, 2029, 2030]

# Column name aliases (normalised to lowercase)
ALIASES = {
    'codigo': 'code', 'código': 'code',
    'titulo': 'title', 'título': 'title', 'titre': 'title',
    'ministerio': 'ministry', 'ministério': 'ministry', 'ministere': 'ministry', 'ministère': 'ministry',
    'sector': 'sector', 'secteur': 'sector', 'setor': 'sector',
    'pilar': 'pillar', 'pilier': 'pillar',
    'financiador': 'financier', 'financeur': 'financier',
    'estado': 'status', 'statut': 'status', 'état': 'status',
    'descricao': 'description', 'descrição': 'description',
    'data_inicio': 'start_date', 'data_fim': 'end_date',
    'regiao': 'region', 'região': 'region',
    'nacional': 'is_national',
}


def _to_decimal(val):
    if val is None or val == '':
        return Decimal('0')
    try:
        # Handle comma as decimal separator
        s = str(val).replace(' ', '').replace(',', '.')
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _normalise_col(name):
    """Normalise column name to canonical key."""
    n = name.strip().lower().replace(' ', '_')
    return ALIASES.get(n, n)


def _resolve_fk(model, value, fields):
    """Try to resolve a FK by multiple fields."""
    if not value:
        return None
    value_str = str(value).strip()
    for field in fields:
        try:
            return model.objects.get(**{field: value_str})
        except model.DoesNotExist:
            continue
        except model.MultipleObjectsReturned:
            return model.objects.filter(**{field: value_str}).first()
    # Case-insensitive fallback
    for field in fields:
        try:
            return model.objects.get(**{f'{field}__iexact': value_str})
        except model.DoesNotExist:
            continue
        except model.MultipleObjectsReturned:
            return model.objects.filter(**{f'{field}__iexact': value_str}).first()
    return None


class Command(BaseCommand):
    help = 'Importa projectos via CSV, Excel (.xlsx) ou JSON.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', type=str, required=True,
            help='Fichier à importer (.csv, .xlsx, .json)',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Analyse sans écrire en base.',
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Supprime les programmations existantes (version=1) avant import.',
        )
        parser.add_argument(
            '--sheet', type=str, default=None,
            help='Nom de la feuille Excel (par défaut: première feuille).',
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.errors = []
        self.stats = {
            'rows_read': 0,
            'projects_created': 0,
            'projects_updated': 0,
            'programming_upserted': 0,
            'rows_skipped': 0,
        }

        file_path = Path(options['file'])
        if not file_path.exists():
            raise CommandError(f'Fichier introuvable : {file_path}')

        ext = file_path.suffix.lower()
        if ext == '.csv':
            rows = self._read_csv(file_path)
        elif ext in ('.xlsx', '.xls'):
            rows = self._read_excel(file_path, options.get('sheet'))
        elif ext == '.json':
            rows = self._read_json(file_path)
        else:
            raise CommandError(f'Format non supporté : {ext}. Utilisez .csv, .xlsx ou .json')

        if not rows:
            raise CommandError('Aucune donnée trouvée dans le fichier.')

        if self.dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] Aucune écriture en base.'))

        self.stdout.write(f'Lignes à traiter : {len(rows)}')

        dgp_tenant = Tenant.objects.filter(is_dgp=True).first()

        with transaction.atomic():
            if self.dry_run:
                from django.db import connection as _conn
                sp = _conn.savepoint()

            if options['clear'] and not self.dry_run:
                deleted, _ = AnnualProgramming.objects.filter(version=1).delete()
                self.stdout.write(self.style.WARNING(
                    f'Programmations supprimées : {deleted}'
                ))

            for i, row in enumerate(rows, start=1):
                self._process_row(i, row, dgp_tenant)

            if self.dry_run:
                _conn.savepoint_rollback(sp)

        # Report
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 50))
        self.stdout.write(self.style.SUCCESS('RAPPORT D\'IMPORT'))
        self.stdout.write(self.style.SUCCESS('=' * 50))
        for k, v in self.stats.items():
            self.stdout.write(f'  {k}: {v}')

        if self.errors:
            self.stdout.write('')
            self.stdout.write(self.style.ERROR(f'Erreurs ({len(self.errors)}) :'))
            for e in self.errors[:50]:
                self.stdout.write(self.style.ERROR(f'  {e}'))
            if len(self.errors) > 50:
                self.stdout.write(self.style.ERROR(
                    f'  ... et {len(self.errors) - 50} autres erreurs.'
                ))

    # ------------------------------------------------------------------
    # Readers
    # ------------------------------------------------------------------

    def _read_csv(self, path):
        rows = []
        with open(path, newline='', encoding='utf-8-sig') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(f, dialect=dialect)
            for raw_row in reader:
                row = {_normalise_col(k): v for k, v in raw_row.items() if k}
                rows.append(row)
        return rows

    def _read_excel(self, path, sheet_name=None):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl non installé. Exécutez : pip install openpyxl')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        rows_iter = ws.iter_rows(values_only=True)
        header_raw = next(rows_iter, None)
        if not header_raw:
            return []

        header = [_normalise_col(str(h or '')) for h in header_raw]
        rows = []
        for vals in rows_iter:
            row = {}
            for col_name, val in zip(header, vals):
                if col_name:
                    row[col_name] = val
            if row.get('code') or row.get('title'):
                rows.append(row)
        wb.close()
        return rows

    def _read_json(self, path):
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, list):
            return [{_normalise_col(k): v for k, v in item.items()} for item in data]
        if isinstance(data, dict) and 'projects' in data:
            return [{_normalise_col(k): v for k, v in item.items()} for item in data['projects']]
        raise CommandError('Format JSON invalide. Attendu : liste ou {"projects": [...]}')

    # ------------------------------------------------------------------
    # Row processing
    # ------------------------------------------------------------------

    def _process_row(self, line_num, row, dgp_tenant):
        self.stats['rows_read'] += 1
        code = str(row.get('code', '')).strip()
        title = str(row.get('title', '')).strip()

        if not code or not title:
            self.errors.append(f'Ligne {line_num}: code ou title manquant — ignorée.')
            self.stats['rows_skipped'] += 1
            return

        # Resolve FKs
        ministry = _resolve_fk(Ministry, row.get('ministry'), ['short_name', 'name'])
        if not ministry and row.get('ministry'):
            self.errors.append(
                f'Ligne {line_num} ({code}): ministère "{row["ministry"]}" non trouvé — ignorée.'
            )
            self.stats['rows_skipped'] += 1
            return

        financier = _resolve_fk(Financier, row.get('financier'), ['short_name', 'name'])
        if not financier and row.get('financier'):
            self.errors.append(
                f'Ligne {line_num} ({code}): financeur "{row["financier"]}" non trouvé — ignorée.'
            )
            self.stats['rows_skipped'] += 1
            return

        sector = _resolve_fk(Sector, row.get('sector'), ['code', 'label'])
        pillar = _resolve_fk(Pillar, row.get('pillar'), ['code', 'label'])
        region = _resolve_fk(Region, row.get('region'), ['name', 'code'])

        # Status
        status_raw = str(row.get('status', 'IDENTIFIED')).strip().upper()
        valid_statuses = [s.value for s in ProjectStatus]
        status = status_raw if status_raw in valid_statuses else ProjectStatus.IDENTIFIED

        # Dates
        start_date = row.get('start_date') or None
        end_date = row.get('end_date') or None

        # is_national
        is_nat = row.get('is_national')
        if isinstance(is_nat, str):
            is_national = is_nat.lower() in ('true', '1', 'sim', 'oui', 'yes')
        elif isinstance(is_nat, bool):
            is_national = is_nat
        else:
            is_national = True

        # Create or update project
        defaults = {
            'title': title,
            'description': str(row.get('description', '') or ''),
            'status': status,
            'is_national': is_national,
        }
        if ministry:
            defaults['ministry'] = ministry
        if financier:
            defaults['principal_financier'] = financier
        if sector:
            defaults['sector'] = sector
        if pillar:
            defaults['pillar'] = pillar
        if region:
            defaults['region'] = region
        if start_date:
            defaults['start_date'] = start_date
        if end_date:
            defaults['end_date'] = end_date
        if dgp_tenant:
            defaults['tenant'] = dgp_tenant

        try:
            proj, created = Project.objects.update_or_create(
                code=code, is_deleted=False, defaults=defaults
            )
        except Exception as e:
            self.errors.append(f'Ligne {line_num} ({code}): erreur projet — {e}')
            self.stats['rows_skipped'] += 1
            return

        if created:
            self.stats['projects_created'] += 1
        else:
            self.stats['projects_updated'] += 1

        # Annual programming
        for yr in ANOS:
            don = _to_decimal(row.get(f'don_{yr}'))
            emp = _to_decimal(row.get(f'emp_{yr}'))
            state = _to_decimal(row.get(f'state_{yr}'))

            if don == 0 and emp == 0 and state == 0:
                continue

            try:
                AnnualProgramming.objects.update_or_create(
                    project=proj, fiscal_year=yr, version=1,
                    defaults={
                        'donations': don,
                        'loans': emp,
                        'state_contribution': state,
                    },
                )
                self.stats['programming_upserted'] += 1
            except Exception as e:
                self.errors.append(
                    f'Ligne {line_num} ({code}): erreur programmation {yr} — {e}'
                )
