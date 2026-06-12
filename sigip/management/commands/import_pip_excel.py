"""
Management command: import_pip_excel
Importa os dados reais do PIP 2026-2030 a partir do ficheiro Excel oficial.

Ficheiro fonte:
  PIP_2026_2030_DPDEP_V2 DGP_5_JUNHO_2026 PIP FINAL.xlsx
  Folha: «GLOBAL FINANÇAS » (com espaço a seguir)

Colunas por ano (índices 0-based):
  2026: D=3 (Donativos), E=4 (Empréstimos), G=6 (Fin.Interno), H=7 (Total)
  2027: I=8, J=9, L=11, M=12
  2028: N=13, O=14, Q=16, R=17
  2029: S=18, T=19, V=21, W=22
  2030: X=23, Y=24, AA=26, AB=27
  Coluna AC=28 : Total Geral PND

Detecção do ministério: linha onde col A é vazia e col B tem texto
(não começa por «TOTAL»).

Uso:
  python manage.py import_pip_excel --file /caminho/ficheiro.xlsx
  python manage.py import_pip_excel  # usa DEFAULT_EXCEL_PATH
  python manage.py import_pip_excel --dry-run
  python manage.py import_pip_excel --clear   # apaga AnnualProgramming antes
"""

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from sigip.models import (
    Pillar, Sector, Ministry, Financier, FinancierType,
    Project, ProjectStatus, WorkflowStatus, AnnualProgramming,
    StateFunction,
)
from core.models import Tenant


# ---------------------------------------------------------------------------
# Column layout (0-indexed)
# ---------------------------------------------------------------------------

YEAR_COLS = {
    # year: (don_idx, emp_idx, interno_idx, total_idx)
    2026: (3, 4, 6, 7),
    2027: (8, 9, 11, 12),
    2028: (13, 14, 16, 17),
    2029: (18, 19, 21, 22),
    2030: (23, 24, 26, 27),
}
GRAND_TOTAL_IDX = 28   # col AC
MAIN_SHEET = 'GLOBAL FINANÇAS '   # trailing space is intentional

REFERENCE_TOTALS = {
    # Reference values (milliers FCFA) from official Excel – used for QC
    2026: Decimal('83200000.00'),
    2027: Decimal('102051496.24'),
    2028: Decimal('115062041.50'),
    2029: Decimal('156727691.40'),
    2030: Decimal('174157691.40'),
    'grand': Decimal('631198919.55'),
}
TOLERANCE = Decimal('1.00')   # max deviation allowed (milliers FCFA)


# ---------------------------------------------------------------------------
# Normalisation maps
# ---------------------------------------------------------------------------

MINISTRY_NORMALISATION = {
    # Excel name → canonical DB name
    'ASSENBLEIA NACIONALPPUILAR':                     'ASSEMBLEIA NACIONAL POPULAR',
    'ASSENBLEIA NACIONAL POPULAR':                    'ASSEMBLEIA NACIONAL POPULAR',
    'ASSEMBLEIA NACIONAL POPULAR':                    'ASSEMBLEIA NACIONAL POPULAR',
    'MINISTERIO DA  COMUNICAÇÃO SOCIAL':              'MINISTÉRIO DA COMUNICAÇÃO SOCIAL',
    'MINISTERIO DA ECONOMIA, PLANO E INTEGRAÇÃO REGIONAL REGIONAL':
        'MINISTÉRIO DA ECONOMIA, PLANO E INTEGRAÇÃO REGIONAL',
    'MINISTERIO DO  AMBIENTE, BIODIVERSIDADE E AÇÃO CLIMATICA':
        'MINISTÉRIO DO AMBIENTE, BIODIVERSIDADE E ACÇÃO CLIMÁTICA',
    'MINISTERIO DO TURISMO E ARTESANATO':             'MINISTÉRIO DO TURISMO E ARTESANATO',
    'MINISTERIO DOS  COMBATENTES DA LIBERDADE DA PATRIA':
        'MINISTÉRIO DOS COMBATENTES DA LIBERDADE DA PÁTRIA',
    'MINISTÉRIO  DA CULTURA , JUVENTUDE   E DESPORTOS':
        'MINISTÉRIO DA CULTURA, JUVENTUDE E DESPORTOS',
    'MINISTÉRIO DA  MULHER, FAMILIA E SOLIDARIEDADE SOCIAL':
        'MINISTÉRIO DA MULHER, FAMÍLIA E SOLIDARIEDADE SOCIAL',
    'COMÉRCIO':                                       'MINISTÉRIO DO COMÉRCIO E INDÚSTRIA',
    'INDÚSTRIA':                                      'MINISTÉRIO DO COMÉRCIO E INDÚSTRIA',
    'ÁGUAS E SANEAMENTO':                             'MINISTÉRIO DAS ÁGUAS E SANEAMENTO',
    'MINISTÉRIO DAS ÁGUAS E SANEAMENTO':              'MINISTÉRIO DAS ÁGUAS E SANEAMENTO',
    'MINISTÉRIO DOS RECURSOS NATURAIS':               'MINISTÉRIO DOS RECURSOS NATURAIS',
}

MINISTRY_SHORT_NAMES = {
    'ASSEMBLEIA NACIONAL POPULAR':                                        'ANP',
    'MINISTÉRIO DA PRESIDENCIA DO CONSELHO DE MINISTROS':                 'Presidência CM',
    'MINISTÉRIO DA COMUNICAÇÃO SOCIAL':                                   'Comunicação',
    'MINISTÉRIO DA ECONOMIA, PLANO E INTEGRAÇÃO REGIONAL':                'Economia/Plano',
    'MINISTÉRIO DO AMBIENTE, BIODIVERSIDADE E ACÇÃO CLIMÁTICA':           'Ambiente',
    'MINISTÉRIO DO TURISMO E ARTESANATO':                                 'Turismo',
    'MINISTÉRIO DOS COMBATENTES DA LIBERDADE DA PÁTRIA':                  'Combatentes',
    'MINISTÉRIO DA CULTURA, JUVENTUDE E DESPORTOS':                       'Cultura/Juventude',
    'MINISTÉRIO DA MULHER, FAMÍLIA E SOLIDARIEDADE SOCIAL':               'Solidariedade',
    'MINISTÉRIO DA ADMINISTRAÇÃO PÚBLICA, TRABALHO, EMPREGO, FORMAÇÃO PROFISSIONAL E SEGURANÇA SOCIAL':
        'Adm. Pública',
    'MINISTÉRIO DA ADMINISTRAÇÃO TERRITORIAL E DO PODER LOCAL':           'Adm. Territorial',
    'MINISTÉRIO DA AGRICULTURA, FLORESTA E DESENVOLVIMENTO RURAL':        'Agricultura',
    'MINISTÉRIO DA DEFESA NACIONAL':                                      'Defesa',
    'MINISTÉRIO DA EDUCAÇÃO NACIONAL, ENSINO SUPERIOR E INVESTIGAÇÃO CIENTIFICA':
        'Educação',
    'MINISTÉRIO DA ENERGIA':                                              'Energia',
    'MINISTÉRIO DA JUSTIÇA E DIREITOS HUMANOS':                           'Justiça',
    'MINISTÉRIO DA SAÚDE PUBLICA':                                        'Saúde',
    'MINISTÉRIO DAS FINANÇAS':                                            'Finanças',
    'MINISTÉRIO DAS OBRAS PÚBLICAS, HABITAÇÃO E URBANISMO':               'Obras Públicas',
    'MINISTÉRIO DAS PESCAS E ECONOMIA MARITIMA':                          'Pescas',
    'MINISTÉRIO DO INTERIOR E DA ORDEM PUBLICA':                          'Interior',
    'MINISTÉRIO DOS TRANSPORTES  TELECOMUNICAÇÃO E ECONOMIA DIGITAL':     'Transportes',
    'MINISTÉRIO DAS ÁGUAS E SANEAMENTO':                                  'Águas/Saneamento',
    'MINISTÉRIO DO COMÉRCIO E INDÚSTRIA':                                 'Comércio/Indústria',
    'MINISTÉRIO DOS RECURSOS NATURAIS':                                   'Recursos Naturais',
}

# Maps raw Excel financier string → canonical key used in DB (Financier.name)
FINANCIER_NORMALISATION = {
    'GOV.GB':   'Estado',
    'GOV. GB':  'Estado',
    'GOV.GB ':  'Estado',
    'JAPÃO':    'JAPAO',
    'GEF-PNUD': 'GEF',
    'GEF7':     'GEF',
    'U.E/FEC':  'EU',
    'U.E./FNUAP': 'EU',
    'FNUAP/UJAB44': 'FNUAP',
    'FNUAP/UJAB93': 'FNUAP',
    'UNFPA/ FM': 'UNFPA',
    'UNFPA/ UNICEF': 'UNFPA',
    'PBF/FAO':  'FAO',
    'MRC/FAO':  'FAO',
    'PAM/FAO':  'FAO',
    'GEF/PNUD': 'GEF',
    'GEF/FNUA/GOV. GB': 'GEF',
    'GEF/BOAD': 'GEF',
    'FVC/PNUD': 'PNUD',
    'FVC/FAO':  'FAO',
    'FVC/OSS/ADPP': 'PNUD',
    'FVC/ECOWAS': 'CEDEAO',
    'FVC/UCGL': 'PNUD',
    'BAD/CILS': 'BAD',
    'BAD/FAD':  'BAD',
    'BAD/CLSS': 'BAD',
    'BAD/FAD/FAT': 'BAD',
    'BAD/BID':  'BAD',
    'BM/JICA':  'BM',
    'BM/BEI':   'BM',
    'BOAD/UEMOA': 'BOAD',
    'BOAD/F.A.Clim': 'BOAD',
    'CEDEAO/GOV. ESPANHA': 'CEDEAO',
    'CEDEAO/CCDG': 'CEDEAO',
    'SIST.NU/UNICEF': 'UNICEF',
    'Sist.NU/UNICEF': 'UNICEF',
    'Sist.NU/HCR/PPG1SENG': 'UNICEF',
    'UNICEF/PAM/PLAN': 'UNICEF',
    'PNUD/UNICEF': 'PNUD',
    'COP.PT/INST.CAM.': 'PORTUGAL',
    'COOP.PORT.': 'PORTUGAL',
    'COOP. PORT': 'PORTUGAL',
    'COP.PT': 'PORTUGAL',
    'ESPANHA/RUSIA': 'ESPANHA',
    'FNUAP/PNUD': 'FNUAP',
    'PROCURA': 'Estado',
    'SOS ALEMANHA': 'Estado',
    'SOS HOLANDA': 'Estado',
    'GOVERNO ALEMAO': 'Estado',
    'SIGHTSAVERS': 'Estado',
    'USAID/PLAN': 'Estado',
    'EUA.SOUTAMTHONSON': 'Estado',
    'PLAN INBTERNATIONAL': 'Estado',
    'UNIVERSITE EXETER': 'Estado',
    'PRCM': 'Estado',
    'MAVA/TARTARUGA': 'Estado',
    'CNO SPAD': 'Estado',
}

FINANCIER_TYPE_MAP = {
    'Estado':        (FinancierType.ESTADO,       'Guiné-Bissau'),
    'BAD':           (FinancierType.MULTILATERAL,  "Côte d'Ivoire"),
    'BM':            (FinancierType.MULTILATERAL,  'États-Unis'),
    'BID':           (FinancierType.MULTILATERAL,  'États-Unis'),
    'BOAD':          (FinancierType.MULTILATERAL,  'Togo'),
    'BADEA':         (FinancierType.MULTILATERAL,  'Arabie Saoudite'),
    'UEMOA':         (FinancierType.MULTILATERAL,  'Sénégal'),
    'CEDEAO':        (FinancierType.MULTILATERAL,  'Nigeria'),
    'FIDA':          (FinancierType.UN_AGENCY,     'Italie'),
    'FAO':           (FinancierType.UN_AGENCY,     'Italie'),
    'PNUD':          (FinancierType.UN_AGENCY,     'États-Unis'),
    'UNICEF':        (FinancierType.UN_AGENCY,     'États-Unis'),
    'PAM':           (FinancierType.UN_AGENCY,     'Italie'),
    'OMS':           (FinancierType.UN_AGENCY,     'Suisse'),
    'FNUAP':         (FinancierType.UN_AGENCY,     'États-Unis'),
    'UNFPA':         (FinancierType.UN_AGENCY,     'États-Unis'),
    'GEF':           (FinancierType.MULTILATERAL,  'États-Unis'),
    'GAVI':          (FinancierType.MULTILATERAL,  'Suisse'),
    'EU':            (FinancierType.MULTILATERAL,  'Belgique'),
    'U.E':           (FinancierType.MULTILATERAL,  'Belgique'),
    'CHINA':         (FinancierType.BILATERAL,     'Chine'),
    'BRASIL':        (FinancierType.BILATERAL,     'Brésil'),
    'JAPAO':         (FinancierType.BILATERAL,     'Japon'),
    'PORTUGAL':      (FinancierType.BILATERAL,     'Portugal'),
    'ESPANHA':       (FinancierType.BILATERAL,     'Espagne'),
    'FM':            (FinancierType.MULTILATERAL,  'États-Unis'),
}

# Sector code (first 2 digits of project code) → StateFunction code
SECTOR_TO_SF = {
    '11': 'SF01',  '12': 'SF02',  '13': 'SF03',
    '21': 'SF09',  '22': 'SF07',  '23': 'SF06',
    '24': 'SF06',  '25': 'SF10',  '26': 'SF08',
    '28': 'SF03',  '31': 'SF04',  '32': 'SF04',
    '33': 'SF04',  '34': 'SF04',  '35': 'SF04',
    '36': 'SF01',  '41': 'SF04',  '42': 'SF04',
    '43': 'SF04',  '44': 'SF04',  '51': 'SF05',
    '52': 'SF05',  '53': 'SF06',  '61': 'SF07',
    '62': 'SF01',  '71': 'SF09',  '81': 'SF01',
    '82': 'SF03',  '99': 'SF04',
}

STATE_FUNCTIONS = [
    ('SF01', 'Serviços Públicos Gerais', 1),
    ('SF02', 'Defesa', 2),
    ('SF03', 'Ordem e Segurança Pública', 3),
    ('SF04', 'Assuntos Económicos', 4),
    ('SF05', 'Protecção do Ambiente', 5),
    ('SF06', 'Habitação e Equipamentos Colectivos', 6),
    ('SF07', 'Saúde', 7),
    ('SF08', 'Lazer, Cultura e Religião', 8),
    ('SF09', 'Educação', 9),
    ('SF10', 'Protecção Social', 10),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_9DIGIT = re.compile(r'^\d{9}$')


def _is_project_code(val):
    return bool(_9DIGIT.match(str(val or '').strip()))


def _to_decimal(val):
    try:
        return Decimal(str(val or 0))
    except InvalidOperation:
        return Decimal('0')


def _normalise_ministry(raw: str) -> str:
    raw = re.sub(r'\s+', ' ', raw).strip()
    return MINISTRY_NORMALISATION.get(raw, raw)


def _normalise_financier(raw: str) -> str:
    raw = raw.strip()
    if raw in FINANCIER_NORMALISATION:
        return FINANCIER_NORMALISATION[raw]
    # Take first component before '/'
    first = raw.split('/')[0].strip()
    if first in FINANCIER_NORMALISATION:
        return FINANCIER_NORMALISATION[first]
    return first


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = (
        'Importa o PIP 2026-2030 a partir do ficheiro Excel oficial (GLOBAL FINANÇAS). '
        'Idempotente: pode ser executado várias vezes. '
        'Actualiza AnnualProgramming com os valores exactos do Excel.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', type=str, default=None,
            help='Caminho para o ficheiro Excel (xlsx).',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Analisa sem gravar na base de dados.',
        )
        parser.add_argument(
            '--clear', action='store_true',
            help='Apaga AnnualProgramming existente antes de importar.',
        )
        parser.add_argument(
            '--skip-qc', action='store_true',
            help='Não aborta mesmo se os totais não correspondem aos valores de referência.',
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.skip_qc = options['skip_qc']
        self.anomalies = []
        self.stats = {
            'rows_read': 0,
            'projects_created': 0, 'projects_updated': 0,
            'programming_created': 0, 'programming_updated': 0,
            'financiers_created': 0, 'ministries_created': 0,
            'anomalies': 0,
        }

        if self.dry_run:
            self.stdout.write(self.style.WARNING('[DRY-RUN] Nada será gravado.'))

        # ---- Locate file ----
        excel_path = options['file']
        if not excel_path:
            # Default: look alongside management commands dir
            excel_path = (
                Path(__file__).resolve().parent.parent.parent.parent
                / 'scripts'
                / 'PIP_2026_2030_DPDEP_V2 DGP_5_JUNHO_2026 PIP FINAL.xlsx'
            )
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise CommandError(
                f'Ficheiro não encontrado: {excel_path}\n'
                f'Use --file /caminho/para/ficheiro.xlsx'
            )

        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl não instalado. Execute: pip install openpyxl')

        self.stdout.write(f'A ler: {excel_path}')
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        if MAIN_SHEET not in wb.sheetnames:
            available = ', '.join(wb.sheetnames)
            raise CommandError(
                f'Folha "{MAIN_SHEET}" não encontrada. Folhas disponíveis: {available}'
            )

        ws = wb[MAIN_SHEET]

        with transaction.atomic():
            if self.dry_run:
                from django.db import connection as _conn
                sp = _conn.savepoint()

            self._ensure_state_functions()
            state_fn_map = {sf.code: sf for sf in StateFunction.objects.all()}

            if options['clear'] and not self.dry_run:
                deleted, _ = AnnualProgramming.objects.filter(version=1).delete()
                self.stdout.write(self.style.WARNING(
                    f'AnnualProgramming v1 apagados: {deleted}'
                ))

            dgp_tenant = Tenant.objects.filter(is_dgp=True).first()
            rows_data = self._parse_sheet(ws)

            calc_totals = {yr: Decimal('0') for yr in range(2026, 2031)}
            calc_grand = Decimal('0')

            for entry in rows_data:
                self.stats['rows_read'] += 1
                proj_obj, fin_obj = self._get_or_create_project(
                    entry, dgp_tenant, state_fn_map
                )
                for yr in range(2026, 2031):
                    don = entry['years'][yr]['donations']
                    emp = entry['years'][yr]['loans']
                    interno = entry['years'][yr]['estado']
                    row_total = don + emp + interno
                    calc_totals[yr] += row_total

                    prog, created = AnnualProgramming.objects.update_or_create(
                        project=proj_obj,
                        fiscal_year=yr,
                        version=1,
                        defaults={
                            'donations': don,
                            'loans': emp,
                            'state_contribution': interno,
                        },
                    )
                    if created:
                        self.stats['programming_created'] += 1
                    else:
                        self.stats['programming_updated'] += 1

                calc_grand += _to_decimal(entry.get('grand_total', 0))

            if not self.dry_run:
                self._run_qc(calc_totals, calc_grand)

            if self.dry_run:
                from django.db import connection as _conn
                _conn.savepoint_rollback(sp)

        self._print_report(calc_totals, calc_grand)

    # -----------------------------------------------------------------------
    # State functions
    # -----------------------------------------------------------------------

    def _ensure_state_functions(self):
        for code, label, order in STATE_FUNCTIONS:
            StateFunction.objects.get_or_create(
                code=code, defaults={'label': label, 'order': order}
            )

    # -----------------------------------------------------------------------
    # Parser
    # -----------------------------------------------------------------------

    def _parse_sheet(self, ws):
        current_ministry = None
        entries = []

        for row in ws.iter_rows(min_row=13, values_only=True):
            col_a = row[0] if len(row) > 0 else None
            col_b = row[1] if len(row) > 1 else None

            a_str = str(col_a or '').strip()
            b_str = str(col_b or '').strip()

            # Ministry header row: col A empty, col B non-empty, not a TOTAL row
            if not a_str and b_str and not b_str.upper().startswith('TOTAL'):
                canonical = _normalise_ministry(b_str)
                if canonical:
                    current_ministry = canonical
                continue

            # Skip non-project rows
            if not _is_project_code(a_str):
                continue

            code = a_str
            title = str(col_b or '').strip() or '(Sem título)'
            fin_raw = str(row[2] or '').strip() if len(row) > 2 else ''
            fin_norm = _normalise_financier(fin_raw)

            years = {}
            for yr, (di, ei, ii, ti) in YEAR_COLS.items():
                don = _to_decimal(row[di] if len(row) > di else 0)
                emp = _to_decimal(row[ei] if len(row) > ei else 0)
                interno = _to_decimal(row[ii] if len(row) > ii else 0)
                total = _to_decimal(row[ti] if len(row) > ti else 0)
                years[yr] = {'donations': don, 'loans': emp, 'estado': interno, 'total': total}

            grand = _to_decimal(row[GRAND_TOTAL_IDX] if len(row) > GRAND_TOTAL_IDX else 0)

            entries.append({
                'code': code,
                'title': title,
                'ministry': current_ministry,
                'financier': fin_norm,
                'years': years,
                'grand_total': grand,
            })

        return entries

    # -----------------------------------------------------------------------
    # Project + Financier
    # -----------------------------------------------------------------------

    def _get_or_create_project(self, entry, dgp_tenant, state_fn_map):
        code = entry['code']
        ministry_name = entry['ministry']
        fin_norm = entry['financier']

        # Ministry
        ministry_obj = None
        if ministry_name:
            short = MINISTRY_SHORT_NAMES.get(ministry_name, '')
            ministry_obj, created = Ministry.objects.get_or_create(
                name=ministry_name,
                defaults={'short_name': short}
            )
            if created:
                self.stats['ministries_created'] += 1
        if not ministry_obj:
            ministry_obj = Ministry.objects.first()
            self._add_anomaly(code, f'Ministério desconhecido para projecto {code}')

        # Financier
        fin_obj = Financier.objects.filter(name=fin_norm).first()
        if not fin_obj:
            ftype, country = FINANCIER_TYPE_MAP.get(fin_norm, (FinancierType.OTHER, ''))
            fin_obj, created = Financier.objects.get_or_create(
                name=fin_norm,
                defaults={'short_name': fin_norm[:50], 'type': ftype, 'country': country}
            )
            if created:
                self.stats['financiers_created'] += 1

        # StateFunction from sector code (first 2 digits)
        sf_code = SECTOR_TO_SF.get(code[:2])
        sf_obj = state_fn_map.get(sf_code) if sf_code else None

        # Sector
        sector_obj = Sector.objects.filter(code=code[:2]).first()

        # Total cost = sum of all year totals
        total_cost = sum(
            entry['years'][yr]['donations'] + entry['years'][yr]['loans'] + entry['years'][yr]['estado']
            for yr in range(2026, 2031)
        )

        project_obj, created = Project.objects.get_or_create(
            code=code,
            defaults={
                'title': entry['title'],
                'ministry': ministry_obj,
                'sector': sector_obj,
                'principal_financier': fin_obj,
                'total_cost': total_cost,
                'status': ProjectStatus.IDENTIFIED,
                'workflow_status': WorkflowStatus.VALIDADO,
                'is_national': True,
                'tenant': dgp_tenant,
                'state_function': sf_obj,
            }
        )

        if created:
            self.stats['projects_created'] += 1
        else:
            # Update fields that may have changed
            updated = False
            if project_obj.ministry != ministry_obj:
                project_obj.ministry = ministry_obj
                updated = True
            if project_obj.total_cost != total_cost:
                project_obj.total_cost = total_cost
                updated = True
            if project_obj.state_function != sf_obj and sf_obj is not None:
                project_obj.state_function = sf_obj
                updated = True
            if updated:
                project_obj.save()
                self.stats['projects_updated'] += 1

        return project_obj, fin_obj

    # -----------------------------------------------------------------------
    # Quality control
    # -----------------------------------------------------------------------

    def _run_qc(self, calc_totals, calc_grand):
        ok = True
        for yr in range(2026, 2031):
            ref = REFERENCE_TOTALS[yr]
            got = calc_totals[yr]
            diff = abs(got - ref)
            if diff > TOLERANCE:
                self.stdout.write(self.style.ERROR(
                    f'  QC FAIL {yr}: calculado={got:.2f}, referência={ref:.2f}, '
                    f'diferença={diff:.2f} (tolerância={TOLERANCE})'
                ))
                ok = False
            else:
                self.stdout.write(self.style.SUCCESS(
                    f'  QC OK   {yr}: {got:.2f} ~= {ref:.2f} (diff={diff:.4f})'
                ))

        ref_grand = REFERENCE_TOTALS['grand']
        diff_grand = abs(calc_grand - ref_grand)
        if diff_grand > TOLERANCE:
            self.stdout.write(self.style.ERROR(
                f'  QC FAIL TOTAL: calculado={calc_grand:.2f}, '
                f'referência={ref_grand:.2f}, diferença={diff_grand:.2f}'
            ))
            ok = False
        else:
            self.stdout.write(self.style.SUCCESS(
                f'  QC OK   TOTAL: {calc_grand:.2f} ~= {ref_grand:.2f} (diff={diff_grand:.4f})'
            ))

        if not ok and not self.skip_qc:
            raise CommandError(
                'QC falhou. Os totais não correspondem aos valores de referência. '
                'Use --skip-qc para ignorar este controlo.'
            )

    # -----------------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------------

    def _add_anomaly(self, code, message):
        self.anomalies.append({'code': code, 'message': message})
        self.stats['anomalies'] += 1

    def _print_report(self, calc_totals, calc_grand):
        self.stdout.write('\n' + '=' * 65)
        self.stdout.write(self.style.SUCCESS('RELATÓRIO IMPORT_PIP_EXCEL – SIGIP-GB 2026-2030'))
        self.stdout.write('=' * 65)
        self.stdout.write(f"  Linhas lidas         : {self.stats['rows_read']}")
        self.stdout.write(f"  Projectos criados    : {self.stats['projects_created']}")
        self.stdout.write(f"  Projectos actualizados: {self.stats['projects_updated']}")
        self.stdout.write(f"  Programações criadas : {self.stats['programming_created']}")
        self.stdout.write(f"  Programações actualizadas: {self.stats['programming_updated']}")
        self.stdout.write(f"  Ministérios criados  : {self.stats['ministries_created']}")
        self.stdout.write(f"  Financiadores criados: {self.stats['financiers_created']}")
        self.stdout.write(f"  Anomalias            : {self.stats['anomalies']}")
        self.stdout.write('')
        self.stdout.write('  Totais calculados (milliers FCFA):')
        for yr in range(2026, 2031):
            self.stdout.write(f'    {yr}: {calc_totals[yr]:.2f}')
        self.stdout.write(f'    TOTAL: {calc_grand:.2f}')

        if self.anomalies:
            self.stdout.write('\n' + self.style.WARNING('ANOMALIAS:'))
            for a in self.anomalies:
                self.stdout.write(f"  {a['code']!r:15s} | {a['message']}")

        self.stdout.write('=' * 65)
        if self.dry_run:
            self.stdout.write(self.style.WARNING('DRY-RUN: nada foi gravado.'))
        else:
            self.stdout.write(self.style.SUCCESS('Importação Excel concluída com sucesso!'))
        self.stdout.write('=' * 65 + '\n')
