"""
SIGIP-GB – Resumo (Summary) Views + Excel Export

4 aggregation endpoints grouping AnnualProgramming data by:
  1. Pillar (PND)
  2. Sector
  3. ExpenseNature (Natureza da Despesa)
  4. GovPriority (Prioridade do Governo)

All grouped by fiscal_year (2026-2030).

Query parameters:
  ?year=2026          – filter to a single fiscal year
  ?validated=true     – include only VALIDADO projects
"""
import io
from decimal import Decimal

from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import AnnualProgramming, StateFunction, ExpenseNature

FISCAL_YEARS = [2026, 2027, 2028, 2029, 2030]

ZERO = Decimal('0')


def _build_response(ordered_rows, grand_total_data, years_to_output):
    """
    Build the final nested JSON response.

    Parameters
    ----------
    ordered_rows : list of (meta_dict, year_map)
        meta_dict: {'id': ..., 'code': ..., 'label': ...}
        year_map:  {fiscal_year: {'donations': Decimal, 'loans': Decimal, 'state': Decimal}}

    grand_total_data : {fiscal_year: {'donations': Decimal, 'loans': Decimal, 'state': Decimal}}

    years_to_output : list[int]
    """
    # ---- Compute grand totals ----
    grand_years = {}
    grand_overall = {'donations': 0.0, 'loans': 0.0, 'state': 0.0,
                     'external': 0.0, 'internal': 0.0, 'total': 0.0}

    for yr in years_to_output:
        d = grand_total_data.get(yr, {})
        don = float(d.get('donations') or 0)
        loan = float(d.get('loans') or 0)
        state = float(d.get('state') or 0)
        ext = don + loan
        total = ext + state
        grand_years[yr] = {
            'donations': don,
            'loans': loan,
            'state': state,
            'external': ext,
            'internal': state,
            'total': total,
        }
        grand_overall['donations'] += don
        grand_overall['loans'] += loan
        grand_overall['state'] += state
        grand_overall['external'] += ext
        grand_overall['internal'] += state
        grand_overall['total'] += total

    grand_years['total'] = grand_overall
    grand_total_overall = grand_overall['total']

    # ---- Build rows ----
    rows = []
    for meta, year_map in ordered_rows:
        row_overall = {'donations': 0.0, 'loans': 0.0, 'state': 0.0,
                       'external': 0.0, 'internal': 0.0, 'total': 0.0, 'pct': 0.0}
        row_years = {}

        for yr in years_to_output:
            d = year_map.get(yr, {})
            don = float(d.get('donations') or 0)
            loan = float(d.get('loans') or 0)
            state = float(d.get('state') or 0)
            ext = don + loan
            total = ext + state
            yr_grand_total = grand_years.get(yr, {}).get('total', 0)
            pct = round(total / yr_grand_total * 100, 4) if yr_grand_total > 0 else 0.0
            row_years[yr] = {
                'donations': don,
                'loans': loan,
                'state': state,
                'external': ext,
                'internal': state,
                'total': total,
                'pct': pct,
            }
            row_overall['donations'] += don
            row_overall['loans'] += loan
            row_overall['state'] += state
            row_overall['external'] += ext
            row_overall['internal'] += state
            row_overall['total'] += total

        row_overall['pct'] = round(
            row_overall['total'] / grand_total_overall * 100, 4
        ) if grand_total_overall > 0 else 0.0
        row_years['total'] = row_overall

        row_dict = dict(meta)
        row_dict['years'] = row_years
        rows.append(row_dict)

    return {
        'rows': rows,
        'grand_total': {'years': grand_years},
    }


def _get_base_qs(request):
    """Apply optional filters from query params."""
    qs = AnnualProgramming.objects.filter(
        project__is_deleted=False,
        version=1,
    )
    validated = request.query_params.get('validated', '').lower()
    if validated in ('true', '1', 'yes'):
        qs = qs.filter(project__workflow_status='VALIDADO')
    year = request.query_params.get('year')
    if year:
        try:
            qs = qs.filter(fiscal_year=int(year))
        except ValueError:
            pass
    return qs


def _parse_years(request):
    """Return the list of years to output based on ?year= param."""
    year_filter = request.query_params.get('year')
    if year_filter:
        try:
            return [int(year_filter)]
        except ValueError:
            pass
    return FISCAL_YEARS


# ---------------------------------------------------------------------------
# 1. Resumo by Pillar (PND)
# ---------------------------------------------------------------------------

class ResumoPNDView(APIView):
    """
    GET /api/v1/resumo/pnd/
    Aggregates AnnualProgramming by Pillar and fiscal_year.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__pillar__id',
                'project__pillar__code',
                'project__pillar__label',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by('project__pillar__order', 'project__pillar__code', 'fiscal_year')
        )

        # Group by pillar key
        # Use tuple as key (hashable); store meta separately
        rows_keys = []       # ordered list of tuple keys
        rows_meta = {}       # key -> meta dict
        rows_data = {}       # key -> {year: {donations, loans, state}}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue
            pid = row['project__pillar__id']
            pcode = row['project__pillar__code'] or ''
            plabel = row['project__pillar__label'] or '(Sem Pilar)'
            key = (pid, pcode, plabel)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': pid, 'code': pcode, 'label': plabel}
                rows_data[key] = {}

            rows_data[key][yr] = {
                'donations': row['donations'],
                'loans': row['loans'],
                'state': row['state'],
            }
            grand_total[yr]['donations'] += row['donations']
            grand_total[yr]['loans'] += row['loans']
            grand_total[yr]['state'] += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[1] or '', k[2] or ''))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


# ---------------------------------------------------------------------------
# 2. Resumo by Sector
# ---------------------------------------------------------------------------

class ResumoSectorView(APIView):
    """
    GET /api/v1/resumo/sector/
    Aggregates AnnualProgramming by Sector and fiscal_year.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__sector__id',
                'project__sector__code',
                'project__sector__label',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by('project__sector__code', 'fiscal_year')
        )

        rows_keys = []
        rows_meta = {}
        rows_data = {}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue
            sid = row['project__sector__id']
            scode = row['project__sector__code'] or ''
            slabel = row['project__sector__label'] or '(Sem Sector)'
            key = (sid, scode, slabel)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': sid, 'code': scode, 'label': slabel}
                rows_data[key] = {}

            rows_data[key][yr] = {
                'donations': row['donations'],
                'loans': row['loans'],
                'state': row['state'],
            }
            grand_total[yr]['donations'] += row['donations']
            grand_total[yr]['loans'] += row['loans']
            grand_total[yr]['state'] += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[1] or '', k[2] or ''))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


# ---------------------------------------------------------------------------
# 3. Resumo by ExpenseNature (Natureza da Despesa)
# ---------------------------------------------------------------------------

class ResumoNaturezaDespesaView(APIView):
    """
    GET /api/v1/resumo/natureza_despesa/
    Aggregates AnnualProgramming by ExpenseNature and fiscal_year.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__expense_nature__id',
                'project__expense_nature__code',
                'project__expense_nature__label',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by('project__expense_nature__code', 'fiscal_year')
        )

        rows_keys = []
        rows_meta = {}
        rows_data = {}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue
            eid = row['project__expense_nature__id']
            ecode = row['project__expense_nature__code'] or ''
            elabel = row['project__expense_nature__label'] or '(Sem Natureza)'
            key = (eid, ecode, elabel)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': eid, 'code': ecode, 'label': elabel}
                rows_data[key] = {}

            rows_data[key][yr] = {
                'donations': row['donations'],
                'loans': row['loans'],
                'state': row['state'],
            }
            grand_total[yr]['donations'] += row['donations']
            grand_total[yr]['loans'] += row['loans']
            grand_total[yr]['state'] += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[1] or '', k[2] or ''))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


# ---------------------------------------------------------------------------
# 4. Resumo by GovPriority (Prioridade do Governo)
# ---------------------------------------------------------------------------

class ResumoPrioridadeGovernoView(APIView):
    """
    GET /api/v1/resumo/prioridade_governo/
    Aggregates AnnualProgramming by GovPriority and fiscal_year.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__gov_priority__id',
                'project__gov_priority__label',
                'project__gov_priority__order',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by('project__gov_priority__order', 'project__gov_priority__label', 'fiscal_year')
        )

        rows_keys = []
        rows_meta = {}
        rows_data = {}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue
            gpid = row['project__gov_priority__id']
            gplabel = row['project__gov_priority__label'] or '(Sem Prioridade)'
            gporder = row['project__gov_priority__order'] or 0
            key = (gpid, gplabel, gporder)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': gpid, 'code': str(gporder), 'label': gplabel}
                rows_data[key] = {}

            rows_data[key][yr] = {
                'donations': row['donations'],
                'loans': row['loans'],
                'state': row['state'],
            }
            grand_total[yr]['donations'] += row['donations']
            grand_total[yr]['loans'] += row['loans']
            grand_total[yr]['state'] += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[2], k[1] or ''))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


# ---------------------------------------------------------------------------
# Export — Excel (openpyxl)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 5. Resumo by StateFunction (Função do Estado – COFOG)
# ---------------------------------------------------------------------------

class ResumoFuncaoEstadoView(APIView):
    """
    GET /api/v1/resumo/funcao_estado/
    Aggregates AnnualProgramming by StateFunction (COFOG) and fiscal_year.
    Projects with no state_function are grouped as '(Não Classificado)'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__state_function__id',
                'project__state_function__code',
                'project__state_function__label',
                'project__state_function__order',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by(
                'project__state_function__order',
                'project__state_function__code',
                'fiscal_year',
            )
        )

        rows_keys = []
        rows_meta = {}
        rows_data = {}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue
            sfid    = row['project__state_function__id']
            sfcode  = row['project__state_function__code'] or ''
            sflabel = row['project__state_function__label'] or '(Não Classificado)'
            sforder = row['project__state_function__order'] or 99
            key = (sfid, sfcode, sflabel, sforder)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': sfid, 'code': sfcode, 'label': sflabel}
                rows_data[key] = {}

            rows_data[key][yr] = {
                'donations': row['donations'],
                'loans':     row['loans'],
                'state':     row['state'],
            }
            grand_total[yr]['donations'] += row['donations']
            grand_total[yr]['loans']     += row['loans']
            grand_total[yr]['state']     += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[3], k[1] or ''))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


# ---------------------------------------------------------------------------
# 6. Resumo Funcionamento / Investimento (top-level ExpenseNature)
# ---------------------------------------------------------------------------

class ResumoFuncionamentoInvestimentoView(APIView):
    """
    GET /api/v1/resumo/funcionamento_investimento/
    Aggregates AnnualProgramming by top-level ExpenseNature category
    (FUNC = Funcionamento, INV = Investimento).
    Only shows top-level codes (no dash in code, i.e. FUNC and INV).
    Projects with no expense_nature are grouped as '(Não Classificado)'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        qs = _get_base_qs(request)
        years_to_output = _parse_years(request)

        data = (
            qs.values(
                'project__expense_nature__id',
                'project__expense_nature__code',
                'project__expense_nature__label',
                'fiscal_year',
            )
            .annotate(
                donations=Coalesce(Sum('donations'), Value(ZERO, output_field=DecimalField())),
                loans=Coalesce(Sum('loans'), Value(ZERO, output_field=DecimalField())),
                state=Coalesce(Sum('state_contribution'), Value(ZERO, output_field=DecimalField())),
            )
            .order_by('project__expense_nature__code', 'fiscal_year')
        )

        rows_keys = []
        rows_meta = {}
        rows_data = {}
        grand_total = {yr: {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}
                       for yr in years_to_output}

        for row in data:
            yr = row['fiscal_year']
            if yr not in years_to_output:
                continue

            eid    = row['project__expense_nature__id']
            ecode  = row['project__expense_nature__code'] or ''
            elabel = row['project__expense_nature__label'] or '(Não Classificado)'

            # Group sub-codes under their parent:
            # 'FUNC-BS' → 'FUNC', 'INV-MT' → 'INV'
            if '-' in ecode:
                parent_code = ecode.split('-')[0]
                # Look up top-level label
                parent = ExpenseNature.objects.filter(code=parent_code).first()
                eid    = parent.id   if parent else None
                elabel = parent.label if parent else parent_code
                ecode  = parent_code

            key = (eid, ecode, elabel)

            if key not in rows_meta:
                rows_keys.append(key)
                rows_meta[key] = {'id': eid, 'code': ecode, 'label': elabel}
                rows_data[key] = {}

            if yr not in rows_data[key]:
                rows_data[key][yr] = {'donations': Decimal('0'), 'loans': Decimal('0'), 'state': Decimal('0')}

            rows_data[key][yr]['donations'] += row['donations']
            rows_data[key][yr]['loans']     += row['loans']
            rows_data[key][yr]['state']     += row['state']
            grand_total[yr]['donations']    += row['donations']
            grand_total[yr]['loans']        += row['loans']
            grand_total[yr]['state']        += row['state']

        ordered = sorted(rows_keys, key=lambda k: (k[1] or 'Z'))
        ordered_rows = [(rows_meta[k], rows_data[k]) for k in ordered]
        return Response(_build_response(ordered_rows, grand_total, years_to_output))


TAB_LABELS = {
    'pnd':               'RESUMO PND',
    'sector':            'RESUMO SECTOR',
    'natureza_despesa':  'NATUREZA DESPESA',
    'prioridade_governo':'PRIORIDADE GOVERNO',
}

TAB_VIEWS = {
    'pnd':               ResumoPNDView,
    'sector':            ResumoSectorView,
    'natureza_despesa':  ResumoNaturezaDespesaView,
    'prioridade_governo':ResumoPrioridadeGovernoView,
}


def _fmt_m(v):
    """Return value in millions FCFA, rounded to 1 decimal."""
    return round(float(v or 0) / 1_000_000, 1)


def _build_xlsx(data, tab_label, years):
    """Build an openpyxl workbook from a resumo data dict."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_label[:31]

    # ---- Styles ----
    hdr_font   = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill1  = PatternFill('solid', fgColor='2A3F52')  # dark blue
    hdr_fill2  = PatternFill('solid', fgColor='3A5068')  # mid blue
    total_fill = PatternFill('solid', fgColor='F4F1EA')  # beige
    bold       = Font(bold=True, size=10)
    normal     = Font(size=10)
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right      = Alignment(horizontal='right', vertical='center')
    left       = Alignment(horizontal='left',  vertical='center', wrap_text=True)
    thin       = Side(style='thin', color='AAAAAA')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt    = '#,##0.0'
    pct_fmt    = '0.0"%"'

    rows  = data.get('rows', [])
    grand = data.get('grand_total', {}).get('years', {})

    # ---- Row 1: Republic header ----
    n_data_cols = len(years) * 4 + 4 + 1  # years×4 sub-cols + total×4 + label col
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_data_cols)
    c = ws.cell(1, 1, 'REPÚBLICA DA GUINÉ-BISSAU — MINISTÉRIO DA ECONOMIA E PLANO')
    c.font = Font(bold=True, size=11)
    c.alignment = center

    # ---- Row 2: Table title ----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_data_cols)
    c = ws.cell(2, 1, f'{tab_label} 2026–2030  (Valores em Milhões FCFA)')
    c.font = Font(bold=True, size=11, color='2A3F52')
    c.alignment = center

    # ---- Row 3 blank ----

    # ---- Row 4: year group headers ----
    col = 2
    ws.cell(4, 1, 'Dimensão').font = hdr_font
    ws.cell(4, 1).fill = hdr_fill1
    ws.cell(4, 1).alignment = center
    ws.cell(4, 1).border = border
    ws.row_dimensions[4].height = 22

    for yr in years:
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
        c = ws.cell(4, col, str(yr))
        c.font = hdr_font; c.fill = hdr_fill1; c.alignment = center; c.border = border
        col += 4

    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 3)
    c = ws.cell(4, col, 'TOTAL 2026–2030')
    c.font = hdr_font; c.fill = hdr_fill1; c.alignment = center; c.border = border

    # ---- Row 5: sub-column headers ----
    sub_headers = ['Ext.', 'Int.', 'Total', '%']
    col = 2
    ws.row_dimensions[5].height = 20
    for _ in range(len(years) + 1):
        for h in sub_headers:
            c = ws.cell(5, col, h)
            c.font = hdr_font; c.fill = hdr_fill2; c.alignment = center; c.border = border
            col += 1

    # ---- Data rows ----
    r = 6
    for row in rows:
        c = ws.cell(r, 1, row.get('label', ''))
        c.font = normal; c.alignment = left; c.border = border
        col = 2
        for yr in years:
            yd = row.get('years', {}).get(yr, {})
            for val, fmt in [
                (yd.get('external', 0), num_fmt),
                (yd.get('internal', 0), num_fmt),
                (yd.get('total', 0),    num_fmt),
                (yd.get('pct', 0),      pct_fmt),
            ]:
                c = ws.cell(r, col, _fmt_m(val) if fmt == num_fmt else round(float(val or 0), 1))
                c.font = normal; c.number_format = fmt
                c.alignment = right; c.border = border
                col += 1
        # Row total
        rt = row.get('years', {}).get('total', {})
        for val, fmt in [
            (rt.get('external', 0), num_fmt),
            (rt.get('internal', 0), num_fmt),
            (rt.get('total', 0),    num_fmt),
            (rt.get('pct', 0),      pct_fmt),
        ]:
            c = ws.cell(r, col, _fmt_m(val) if fmt == num_fmt else round(float(val or 0), 1))
            c.font = bold; c.number_format = fmt
            c.alignment = right; c.border = border
            col += 1
        r += 1

    # ---- Grand total row ----
    ws.row_dimensions[r].height = 18
    c = ws.cell(r, 1, 'TOTAL GERAL')
    c.font = bold; c.fill = total_fill; c.alignment = left; c.border = border
    col = 2
    for yr in years:
        gd = grand.get(yr, {})
        for val, fmt in [
            (gd.get('external', 0), num_fmt),
            (gd.get('internal', 0), num_fmt),
            (gd.get('total', 0),    num_fmt),
            (100.0,                 pct_fmt),
        ]:
            c = ws.cell(r, col, _fmt_m(val) if fmt == num_fmt else round(float(val), 1))
            c.font = bold; c.fill = total_fill
            c.number_format = fmt; c.alignment = right; c.border = border
            col += 1
    gt = grand.get('total', {})
    for val, fmt in [
        (gt.get('external', 0), num_fmt),
        (gt.get('internal', 0), num_fmt),
        (gt.get('total', 0),    num_fmt),
        (100.0,                 pct_fmt),
    ]:
        c = ws.cell(r, col, _fmt_m(val) if fmt == num_fmt else round(float(val), 1))
        c.font = bold; c.fill = total_fill
        c.number_format = fmt; c.alignment = right; c.border = border
        col += 1

    # ---- Column widths ----
    ws.column_dimensions['A'].width = 30
    col_letter = 'B'
    from openpyxl.utils import get_column_letter
    for i in range(2, n_data_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    # ---- Freeze header rows ----
    ws.freeze_panes = 'B6'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ResumoExportView(APIView):
    """
    GET /api/v1/resumo/export/?tab=pnd&format=xlsx
    Generates and downloads an Excel file for any of the 4 resumo tabs.

    Query params:
        tab      : pnd | sector | natureza_despesa | prioridade_governo  (default: pnd)
        format   : xlsx  (pdf handled client-side via window.print)
        year     : 2026..2030  (optional, same as resumo views)
        validated: true        (optional)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        tab = request.query_params.get('tab', 'pnd')
        if tab not in TAB_VIEWS:
            return HttpResponse('Invalid tab. Use: pnd, sector, natureza_despesa, prioridade_governo', status=400)

        # Re-use the corresponding view's get() to get the data dict
        view_instance = TAB_VIEWS[tab]()
        view_instance.request = request
        view_instance.args = args
        view_instance.kwargs = kwargs
        response = view_instance.get(request, *args, **kwargs)
        data = response.data

        years = _parse_years(request)
        label = TAB_LABELS[tab]
        buf = _build_xlsx(data, label, years)

        year_suffix = f'_{request.query_params["year"]}' if request.query_params.get('year') else '_2026-2030'
        filename = f'SIGIP_GB_{label.replace(" ", "_")}{year_suffix}.xlsx'

        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
